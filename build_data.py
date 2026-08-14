#!/usr/bin/env python3
"""Extract post + sentiment data and inline it into index.html between the
/*__DATA_START__*/ ... /*__DATA_END__*/ markers.

Posts now come from a CSV export (one row = one post; a creator can span many
rows). Sentiments still come from the second xlsx workbook.

Re-run whenever the source files change:
    python3 build_data.py
"""

from __future__ import annotations

import csv
import json
import re
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
# Updated social export — one row per post, grouped by HANDLE (col D).
POSTS_CSV = Path.home() / "Downloads" / "Aug 12 9am - Sheet1 (1).csv"
# Same data as an .xlsx — CSV drops hyperlink targets, but the xlsx keeps the
# Google Drive deep links behind the Instagram-story cells (col V). We read the
# story hyperlinks from here and use them in place of the CSV's display text.
POSTS_XLSX = Path.home() / "Downloads" / "Aug 12 9am.xlsx"
SENT_XLSX = Path.home() / "Downloads" / "second.xlsx"
# Brooklyn Advance Creator Screening — Instagram only, separate creator roster.
BK_XLSX = Path.home() / "Downloads" / "BK 8.12 10.30am.xlsx"
INDEX_HTML = HERE / "index.html"

EVENT_VENICE = "venice"
EVENT_BROOKLYN = "brooklyn"


# ---------- parsing helpers ----------

FLAT_MENTION_RE = re.compile(r"@([A-Za-z0-9_.]+)")           # Instagram-style flat handles
TT_MENTION_RE = re.compile(r"@([^@#\n\r]+)")                   # TikTok display-name style
HASHTAG_RE = re.compile(r"#([A-Za-z0-9_]+)")
TRAIL_PUNCT_RE = re.compile(r"[^\w\-\s.]+$")
NORMALIZE_RE = re.compile(r"[\s\-._]+")
K_M_RE = re.compile(r"^([\d.]+)\s*([KkMm])$")


def parse_metric(v) -> int | None:
    """Convert numeric or 'K'/'M' string to int; return None for missing/hidden."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip()
    if not s or s.lower() in ("hidden", "-", "—", "n/a", "na"):
        return None
    m = K_M_RE.match(s)
    if m:
        n = float(m.group(1))
        mult = 1000 if m.group(2).lower() == "k" else 1_000_000
        return int(n * mult)
    try:
        return int(float(s.replace(",", "")))
    except ValueError:
        return None


def parse_following(v) -> int | None:
    """TikTok following column — usually a float, sometimes N/A."""
    return parse_metric(v)


def extract_mentions(text: str | None, style: str = "flat") -> list[dict]:
    """Return [{norm, display}] deduped by norm.

    Instagram captions use flat handles (@spidermanmovie) — use `style="flat"`.
    TikTok captions inline display-name tags (@Spider-Man Movie) that can
    contain spaces and hyphens — use `style="tiktok"`. Both normalize to the
    same key (lowercase, separators removed) so they bucket together.
    """
    if not text:
        return []
    seen = {}
    if style == "flat":
        for h in FLAT_MENTION_RE.findall(text):
            display = h
            norm = NORMALIZE_RE.sub("", h).lower()
            if not norm or not norm[0].isalnum():
                continue
            if norm not in seen:
                seen[norm] = display
    else:  # tiktok display-name style
        for raw in TT_MENTION_RE.findall(text):
            chunk = raw.strip()
            chunk = TRAIL_PUNCT_RE.sub("", chunk).strip()
            if not chunk:
                continue
            # take up to first word cluster: word (space word){0,3}, max 40 chars
            short = re.match(r"[\w.\-]+(?:\s+[\w.\-]+){0,3}", chunk)
            if short:
                chunk = short.group(0).strip()
            if len(chunk) > 40:
                chunk = chunk[:40].rsplit(" ", 1)[0]
            norm = NORMALIZE_RE.sub("", chunk).lower()
            if not norm or not norm[0].isalnum():
                continue
            if norm not in seen:
                seen[norm] = chunk
    return [{"norm": k, "display": v} for k, v in seen.items()]


def extract_hashtags(*texts: str | None) -> list[str]:
    seen = {}
    for text in texts:
        if not text:
            continue
        for tag in HASHTAG_RE.findall(text):
            key = tag.lower()
            if key not in seen:
                seen[key] = tag
    return list(seen.values())


def is_valid_url(v) -> bool:
    if not v:
        return False
    s = str(v).strip().lower()
    if s in ("", "no", "yes", "n/a", "na"):
        return False
    return s.startswith("http")


TT_USER_RE = re.compile(r"tiktok\.com/@([A-Za-z0-9_.]+)", re.I)
# instagram profile URLs (not /p/, /reel/, /stories/) → username
IG_USER_RE = re.compile(r"instagram\.com/(?!p/|reel/|stories/|explore/)([A-Za-z0-9_.]+)", re.I)


def tiktok_username(*vals) -> str | None:
    for v in vals:
        if isinstance(v, str):
            m = TT_USER_RE.search(v)
            if m:
                return m.group(1).lower().strip(".")
    return None


def instagram_username(*vals) -> str | None:
    for v in vals:
        if isinstance(v, str):
            m = IG_USER_RE.search(v)
            if m:
                return m.group(1).lower().strip(".")
    return None


# ---------- sentiment category mapping ----------

# Screenshot has: FUN, FAVORITE, STREET LOVE, ICONIC, STOKED, EMOTIONAL, CRYING,
# IMPRESSED, WHOA, HYPED, BUZZING, PLOT TWIST. Assign each quote a category by
# scanning the text for keywords.
SENTIMENT_CATEGORIES = [
    ("PLOT TWIST",  "🤯", "yellow", ["nothing ever happens", "plot twist"]),
    ("WHOA",        "😳", "pink",   ["wait…", "wait is that real", "wait... is that real", "wait is that", "is that real"]),
    ("CRYING",      "🍕", "yellow", ["joe's pizza", "going to cry", "gonna cry"]),
    ("STREET LOVE", "🕷️", "teal",  ["spidey! spidey", "we love you"]),
    ("ICONIC",      "🗽", "yellow", ["ny shit in venice", "some ny shit"]),
    ("STOKED",      "🤩", "pink",   ["dream team", "coolest activation", "can't stop smiling"]),
    ("EMOTIONAL",   "🥹", "teal",   ["dreams coming true", "dream when i first joined"]),
    ("IMPRESSED",   "😍", "teal",   ["really impressed", "so hospitable"]),
    ("FAVORITE",    "🏆", "yellow", ["favorite lighthouse event", "favorite event"]),
    ("HYPED",       "🔥", "pink",   ["insane event", "you killed it", "you're the goat"]),
    # BUZZING = the public "great coverage" moment only. The internal RSVP/reach
    # metrics note is intentionally excluded so the grid matches the curated 12.
    ("BUZZING",     "📈", "teal",   ["great coverage"]),
    ("FUN",         "🎉", "pink",   ["do more stuff like this", "this event is so fun", "so fun"]),
]


def categorize_sentiment(quote: str) -> tuple[str, str, str] | None:
    """Return (label, emoji, tone) if the quote matches one of the 12 curated
    moment categories, or None if it's generic. Callers drop None rows so the
    grid only shows quotable, categorizable moments (matches the design)."""
    q = (quote or "").lower()
    for label, emoji, tone, keywords in SENTIMENT_CATEGORIES:
        for kw in keywords:
            if kw in q:
                return label, emoji, tone
    return None


# ---------- excerpt & attribution shaping for sentiment cards ----------

def shorten_quote(text: str, max_len: int = 200) -> str:
    """Pull the first quoted string if the cell has one; otherwise trim.

    The source cells often prefix context ("Walking inside when Modern Tarzan
    was on the door - \"Wait is that real?\""), which reads poorly on a card.
    """
    t = text.strip()
    m = re.search(r"[\"“]([^\"”]+)[\"”]", t)
    if m:
        return m.group(1).strip()
    if len(t) > max_len:
        return t[: max_len - 1].rstrip() + "…"
    return t


def clean_attribution(name: str) -> tuple[str, str]:
    """Return (name, role) — many attribution cells embed both."""
    n = (name or "").strip()
    # heuristics for the specific dataset
    lookup = {
        "someone walking by windward circle": ("Passerby", "at Windward Circle"),
        "people on the street when we were doing the photoshoot with @moderntarzan":
            ("People on the street", "During the @moderntarzan photoshoot"),
        "member": ("The Lighthouse Member", "Overheard"),
        "member jesse wellens": ("Jesse Wellens", "The Lighthouse Member"),
        "dylan bradshaw and nate norell": ("Dylan Bradshaw & Nate Norell", "Creators"),
    }
    key = n.lower()
    if key in lookup:
        return lookup[key]
    return n, ""


# ---------- main extraction ----------

# The CSV's "Instagram story" cells DISPLAY an instagram.com profile URL, but the
# real target is a Google Drive deep link (the uploaded story recording). CSV
# export flattens the hyperlink to its display text, so those targets are lost.
# These are the deep links recovered from the source workbooks + supplied by the
# client, keyed by lowercased HANDLE (col D). Add more here as they're provided.
STORY_DEEPLINK_OVERRIDES = {
    "thehijabibounder": "https://drive.google.com/drive/folders/1JXvWcJ1aESBbJrEGtv4NGlTTh_G7cFgJ",
    "gabiwhiting": "https://drive.google.com/drive/folders/1DF454zBCa3H7m5oDTjQxPFyxWjfh7SUm",
    "4k.film":     "https://drive.google.com/file/d/1yuWj2b3uN-xURC0rMuR03lsO5h-IwwyN/view?usp=drive_link",
    "_bucketjosh": "https://drive.google.com/file/d/14eurMvrsBiv3CuLamvHdgp6sW5SjhEtm/view?usp=drive_link",
    "alxxpaul":    "https://drive.google.com/file/d/15C7Sv9aI-DIUaSWxwN-rsjAddSmxhhrk/view?usp=drive_link",
    "bakerjayyy":  "https://drive.google.com/file/d/1ewWODWfIrQA3h9F43DkM2CrtIGONREeo/view?usp=sharing",
    "disway80":    "https://drive.google.com/file/d/1HWdo2RENeCsq4vP442YsR1BMY6Iux9jE/view?usp=drive_link",
    "dripxzyt":    "https://drive.google.com/file/d/1eXoH3Er7k6-NtZ_gj0nsDCcmuba-mtsp/view?usp=sharing",
    "kayladsoto":  "https://drive.google.com/file/d/1XXeOf7JGk6a9VxRXDoZtgCVlU-mAtJZt/view?usp=drive_link",
    "nerdcomicsg": "https://drive.google.com/file/d/1evfocMZUvjLnkgSwQg4djfags3yyQdxy/view?usp=sharing",
}


_HYPERLINK_RE = re.compile(r'=HYPERLINK\(\s*"([^"]+)"', re.I)


def load_story_deeplinks() -> dict[str, str]:
    """Read the Instagram-story hyperlink TARGETS (Google Drive deep links) from
    the .xlsx, keyed by lowercased HANDLE. CSV export flattens these to their
    instagram.com display text, so the .xlsx is the only place they survive.

    Handles both real hyperlink objects and =HYPERLINK("url","label") formulas.
    Falls back to the hardcoded overrides if the workbook is missing."""
    if not POSTS_XLSX.exists():
        print(f"NOTE: {POSTS_XLSX.name} not found — using hardcoded story overrides only")
        return dict(STORY_DEEPLINK_OVERRIDES)
    wb = load_workbook(POSTS_XLSX)  # NOT data_only — we need formulas/hyperlinks
    ws = wb[wb.sheetnames[0]]
    out: dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        h = ws.cell(r, 4).value
        h = str(h).strip().lower() if h else ""
        if not h:
            continue
        cell = ws.cell(r, C_IG_STORY + 1)  # openpyxl is 1-based
        target = None
        if cell.hyperlink and cell.hyperlink.target:
            target = cell.hyperlink.target
        elif isinstance(cell.value, str):
            m = _HYPERLINK_RE.search(cell.value)
            if m:
                target = m.group(1)
        if target and "drive.google" in target and h not in out:
            out[h] = target
    # hardcoded overrides fill any gap the workbook doesn't cover
    for h, url in STORY_DEEPLINK_OVERRIDES.items():
        out.setdefault(h, url)
    return out


def _cell(row: list[str], i: int) -> str:
    return (row[i] if i < len(row) else "").strip()


# CSV column indices (0-based) — see header row of the export.
C_CHECKIN, C_FIRST, C_LAST, C_HANDLE = 0, 1, 2, 3
C_TT_FOLLOW, C_TT_LINK, C_TT_POST, C_TT_CAP = 4, 5, 6, 7
C_TT_LIKES, C_TT_COMMENTS, C_TT_SAVES, C_TT_SHARES = 8, 9, 10, 11
C_IG_HANDLE, C_IG_FOLLOW, C_IG_POST, C_IG_CAP, C_IG_HASH = 12, 13, 14, 15, 16
C_IG_LIKES, C_IG_COMMENTS, C_IG_REPOSTS, C_IG_SHARES, C_IG_STORY = 17, 18, 19, 20, 21


def build_posts() -> tuple[list[dict], list[dict], dict]:
    """Read the CSV (one row = one post) and fold it into one record per
    creator (keyed by HANDLE). A creator can own many TikTok and/or Instagram
    posts, plus at most one Instagram story link (col V)."""
    with open(POSTS_CSV, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    data = rows[1:]  # skip header

    creators: dict[str, dict] = {}
    posts: list[dict] = []
    eng = {"likes": 0, "comments": 0, "savesReposts": 0, "shares": 0}
    checked_in: set[str] = set()
    stories_seen: set[str] = set()
    story_deeplinks = load_story_deeplinks()

    for row in data:
        handle = _cell(row, C_HANDLE)
        if not handle:
            continue
        hkey = handle.lower()
        first = _cell(row, C_FIRST)
        last = _cell(row, C_LAST)
        if _cell(row, C_CHECKIN).lower() == "true":
            checked_in.add(hkey)

        tt_following = parse_following(_cell(row, C_TT_FOLLOW))
        ig_following = parse_following(_cell(row, C_IG_FOLLOW))
        tt_profile = _cell(row, C_TT_LINK) or None
        tt_post_url = _cell(row, C_TT_POST)
        tt_caption = _cell(row, C_TT_CAP)
        ig_post_url = _cell(row, C_IG_POST)
        ig_caption = _cell(row, C_IG_CAP)
        ig_hashtags_raw = _cell(row, C_IG_HASH)
        ig_story = _cell(row, C_IG_STORY)

        # ----- creator record (aggregated across the handle's rows) -----
        display_name = " ".join(x for x in (first, last if last and last.upper() != "N/A" else "") if x).strip() or handle
        tt_user = tiktok_username(tt_profile, tt_post_url)
        ig_user = instagram_username(ig_post_url, ig_story)
        creator = creators.setdefault(hkey, {
            "handle": handle,
            "name": display_name,
            "event": EVENT_VENICE,
            "tiktokFollowing": None,
            "igFollowing": None,
            "tiktokProfile": None,
            "ttUser": None,
            "igUser": handle.lstrip("@"),
        })
        # keep the largest follower figures seen for this creator
        if tt_following and tt_following > (creator["tiktokFollowing"] or 0):
            creator["tiktokFollowing"] = tt_following
        if ig_following and ig_following > (creator["igFollowing"] or 0):
            creator["igFollowing"] = ig_following
        if tt_profile and not creator["tiktokProfile"]:
            creator["tiktokProfile"] = tt_profile
        if tt_user and not creator["ttUser"]:
            creator["ttUser"] = tt_user
        if ig_user:
            creator["igUser"] = ig_user
        if display_name and (not creator["name"] or creator["name"] == creator["handle"]):
            creator["name"] = display_name

        # ----- TikTok post -----
        if is_valid_url(tt_post_url):
            tt_likes = parse_metric(_cell(row, C_TT_LIKES))
            tt_comments = parse_metric(_cell(row, C_TT_COMMENTS))
            tt_saves = parse_metric(_cell(row, C_TT_SAVES))
            tt_shares = parse_metric(_cell(row, C_TT_SHARES))
            mentions = extract_mentions(tt_caption, style="tiktok")
            hashtags = extract_hashtags(tt_caption)
            posts.append({
                "creatorHandle": handle,
                "event": EVENT_VENICE,
                "platform": "tiktok",
                "postUrl": tt_post_url,
                "caption": tt_caption or "",
                "mentions": mentions,
                "hashtags": hashtags,
                "metrics": {"likes": tt_likes, "comments": tt_comments,
                            "saves": tt_saves, "shares": tt_shares},
                "tagsCount": len(mentions),
            })
            eng["likes"] += tt_likes or 0
            eng["comments"] += tt_comments or 0
            eng["savesReposts"] += tt_saves or 0
            eng["shares"] += tt_shares or 0

        # ----- Instagram post -----
        if is_valid_url(ig_post_url):
            ig_likes = parse_metric(_cell(row, C_IG_LIKES))
            ig_comments = parse_metric(_cell(row, C_IG_COMMENTS))
            ig_reposts = parse_metric(_cell(row, C_IG_REPOSTS))
            ig_shares = parse_metric(_cell(row, C_IG_SHARES))
            mentions = extract_mentions(ig_caption, style="flat")
            hashtags = extract_hashtags(ig_caption, ig_hashtags_raw)
            posts.append({
                "creatorHandle": handle,
                "event": EVENT_VENICE,
                "platform": "instagram",
                "postUrl": ig_post_url,
                "caption": ig_caption or "",
                "hashtagsExtra": ig_hashtags_raw or "",
                "mentions": mentions,
                "hashtags": hashtags,
                "metrics": {"likes": ig_likes, "comments": ig_comments,
                            "reposts": ig_reposts, "shares": ig_shares},
                "tagsCount": len(mentions),
            })
            eng["likes"] += ig_likes or 0
            eng["comments"] += ig_comments or 0
            eng["savesReposts"] += ig_reposts or 0
            eng["shares"] += ig_shares or 0

        # ----- Instagram STORY — one per creator, linking to col V -----
        if ig_story and hkey not in stories_seen:
            stories_seen.add(hkey)
            urls = [t for t in re.split(r"[,\s]+", ig_story) if t.lower().startswith("http")]
            # Prefer the recovered Google Drive deep link over the CSV's
            # instagram.com display text (which is the wrong target).
            deep = story_deeplinks.get(hkey)
            story_url = deep or (urls[0] if urls else None)
            posts.append({
                "creatorHandle": handle,
                "event": EVENT_VENICE,
                "platform": "story",
                "postUrl": story_url,
                "storyRaw": ig_story,
                "storyCount": len(urls) or 1,
                "isDeepLink": bool(deep),
                "caption": "",
                "mentions": [],
                "hashtags": [],
                "metrics": {},
                "tagsCount": 0,
            })

    # per-creator total following = TikTok following + Instagram followers
    for c in creators.values():
        c["totalFollowing"] = (c.get("tiktokFollowing") or 0) + (c.get("igFollowing") or 0)

    combined_follow = sum(c["totalFollowing"] for c in creators.values())
    million = sum(1 for c in creators.values() if (c.get("tiktokFollowing") or 0) >= 1_000_000)
    tt_count = sum(1 for p in posts if p["platform"] == "tiktok")
    ig_count = sum(1 for p in posts if p["platform"] == "instagram")
    stats = {
        "checkedIn": len(checked_in),
        "creators": len(creators),
        "combinedFollowing": combined_follow,
        "millionPlus": million,
        "postsLogged": tt_count + ig_count,
        "tiktokPosts": tt_count,
        "instagramPosts": ig_count,
        "engagement": eng,
        "totalEngagement": eng["likes"] + eng["comments"] + eng["savesReposts"] + eng["shares"],
    }

    return list(creators.values()), posts, stats


def build_brooklyn() -> tuple[list[dict], list[dict], dict]:
    """Brooklyn Advance Creator Screening — Instagram only.

    The export's Instagram columns are not always in the documented positions:
    continuation rows (a creator's 2nd/3rd post, which carry no follower count)
    are shifted one column to the left. Rather than trusting fixed indices we
    locate the post URL per row and read the remaining fields relative to it,
    which handles both layouts.
    """
    if not BK_XLSX.exists():
        print(f"NOTE: {BK_XLSX.name} not found — skipping Brooklyn")
        return [], [], {}

    wb = load_workbook(BK_XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]

    creators: dict[str, dict] = {}
    posts: list[dict] = []
    eng = {"likes": 0, "comments": 0, "savesReposts": 0, "shares": 0}
    seen_urls: set[str] = set()
    dupes = 0

    for r in range(2, ws.max_row + 1):
        handle = ws.cell(r, 1).value
        handle = str(handle).strip() if handle else ""
        if not handle:
            continue
        hkey = handle.lower()
        name = (str(ws.cell(r, 2).value).strip() if ws.cell(r, 2).value else "") or handle
        tt_following = parse_following(ws.cell(r, 3).value)

        # locate the Instagram post URL (col 13 normally, col 12 on shifted rows)
        url_col = None
        for c in (13, 12):
            if is_valid_url(ws.cell(r, c).value):
                url_col = c
                break

        ig_following = parse_following(ws.cell(r, 12).value) if url_col != 12 else None

        c = creators.setdefault(hkey, {
            "handle": handle,
            "name": name,
            "event": EVENT_BROOKLYN,
            "tiktokFollowing": None,
            "igFollowing": None,
            "tiktokProfile": None,
            "ttUser": None,
            "igUser": (str(ws.cell(r, 11).value).strip().lstrip("@")
                       if ws.cell(r, 11).value else handle.lstrip("@")),
        })
        if tt_following and tt_following > (c["tiktokFollowing"] or 0):
            c["tiktokFollowing"] = tt_following
        if ig_following and ig_following > (c["igFollowing"] or 0):
            c["igFollowing"] = ig_following

        if url_col is None:
            continue                      # attended but didn't post — excluded by request

        post_url = str(ws.cell(r, url_col).value).strip()
        if post_url in seen_urls:
            dupes += 1                    # same URL logged against two creators
        seen_urls.add(post_url)

        caption = ws.cell(r, url_col + 1).value or ""
        caption = str(caption)
        likes = parse_metric(ws.cell(r, url_col + 2).value)
        comments = parse_metric(ws.cell(r, url_col + 3).value)
        reposts = parse_metric(ws.cell(r, url_col + 4).value)
        shares = parse_metric(ws.cell(r, url_col + 5).value)

        mentions = extract_mentions(caption, style="flat")
        hashtags = extract_hashtags(caption)
        posts.append({
            "creatorHandle": handle,
            "event": EVENT_BROOKLYN,
            "platform": "instagram",
            "postUrl": post_url,
            "caption": caption,
            "mentions": mentions,
            "hashtags": hashtags,
            "metrics": {"likes": likes, "comments": comments,
                        "reposts": reposts, "shares": shares},
            "tagsCount": len(mentions),
        })
        eng["likes"] += likes or 0
        eng["comments"] += comments or 0
        eng["savesReposts"] += reposts or 0
        eng["shares"] += shares or 0

    # drop creators who never posted, then total up followings
    posted = {p["creatorHandle"].lower() for p in posts}
    creators = {k: v for k, v in creators.items() if k in posted}
    for c in creators.values():
        c["totalFollowing"] = (c.get("tiktokFollowing") or 0) + (c.get("igFollowing") or 0)

    stats = {
        "creators": len(creators),
        "combinedFollowing": sum(c["totalFollowing"] for c in creators.values()),
        "millionPlus": sum(1 for c in creators.values() if (c.get("tiktokFollowing") or 0) >= 1_000_000),
        "postsLogged": len(posts),
        "tiktokPosts": 0,
        "instagramPosts": len(posts),
        "engagement": eng,
        "totalEngagement": sum(eng.values()),
    }
    if dupes:
        print(f"  ⚠ Brooklyn: {dupes} post URL(s) appear on more than one creator row")
    return list(creators.values()), posts, stats


# ---------- sentiment role buckets ----------
# Every sentiment card's sub-label must read exactly one of these three, so the
# grid stays scannable (it previously showed ad-hoc labels like "Event feedback",
# "Lighthouse feedback", "Sentiment", "During the @moderntarzan photoshoot").
ROLE_CREATOR = "Creator"
ROLE_CLIENT = "Client"
ROLE_OVERHEARD = "Overheard"

# Attributions that are ambient/overheard rather than a named person's feedback.
OVERHEARD_NAMES = {
    "passerby",
    "people on the street",
    "the lighthouse member",
}


def normalize_role(name: str, role: str, source: str) -> str:
    """Collapse the assorted source labels into Creator / Client / Overheard."""
    n = (name or "").strip().lower()
    r = (role or "").strip().lower()
    if source == "client":
        return ROLE_CLIENT
    if n in OVERHEARD_NAMES or "overheard" in r or "passerby" in n:
        return ROLE_OVERHEARD
    return ROLE_CREATOR


# ---------- curated creator quotes pulled from social posts ----------
# (name, handle, quote, category, emoji, tone). `name` is None when the handle
# isn't in the creator database — those cards show the @handle instead.
CURATED_CREATOR_QUOTES = [
    ("_bucketjosh", "I got to see stuff from Spider-Man: Brand New Day that the rest of the world hasn't", "FIRST LOOK", "👀", "yellow"),
    ("therealdoomblazer", "Got to hang with Jackie Chan's stunt team!", "HYPED", "🔥", "pink"),
    ("officialhannahsarah", "How lovely to meet and sit with the editors of brand new day just to chat about our love for this film…just to hear their love for what they do and what we see in film, magic", "CRAFT", "🎬", "teal"),
    ("impatrickt", "Wrapped an insane week in LA, moderating a cinematography panel", "CRAFT", "🎬", "teal"),
    ("jackiebonsignore", "Wait is that @coyjandreau ???", "WHOA", "😳", "pink"),
    ("doublearugs", "Oh my gosh. They had the brand new day suit", "WHOA", "😳", "pink"),
    ("tapeselects", "We were treated to a ton of incredible behind the scenes looks", "FIRST LOOK", "👀", "yellow"),
    ("bakerjayyy", "They played two of my spidey videos!", "STOKED", "🤩", "pink"),
    ("therealdoomblazer", "Long awaited meet up", "FUN", "🎉", "pink"),
    ("therealdoomblazer", "7 year old me would be going crazy right now", "EMOTIONAL", "🥹", "teal"),
    ("minisuperheroestoday", "This was a dream experience", "EMOTIONAL", "🥹", "teal"),
    ("kayladsoto", "Y'all today has been wild. I'm still shook", "HYPED", "🔥", "pink"),
    ("officialhannahsarah", "Things are about to be webtastic today", "FUN", "🎉", "pink"),
    ("macfarlanebros", "Got to hear from the director and heads of departments about their process creating @spidermanmovie !!", "CRAFT", "🎬", "teal"),
    ("joshvstheworld__", "I'm still taking this in brb", "EMOTIONAL", "🥹", "teal"),
    ("dylanjbradshaw", "Got a sneak peek of the new Spider-Man film, met @destindaniel @brettsbo, and more of the incredible team that brought this movie to life!", "FIRST LOOK", "👀", "yellow"),
    ("doublearugs", "Yesterday I had the privilege of attending a Spider-Man: Brand New Day event for creators with Destin Daniel Cretton in attendance, as well as the creative team from the movie… just hearing from the creatives and just knowing that they put so much passion and love into this movie has me excited.", "IMPRESSED", "😍", "teal"),
    ("_bucketjosh", "hearing Destin, the director, along with the rest of the crew talk about the film really put into perspective how much raw passion was put into it.", "IMPRESSED", "😍", "teal"),
    ("lukesoutpost", "Got to hear from the film's director, Destin, and other members of the filmmaking team. Showing us behind the scenes look at how they approached the cinematography…", "CRAFT", "🎬", "teal"),
    ("lukesoutpost", "Getting to hear from people behind the camera was the highlight of the day", "CRAFT", "🎬", "teal"),
    ("disway80", "Thank you for showcasing so many creators from all around the world at todays event.", "GRATEFUL", "🙏", "pink"),
]

# Post URLs for quotes that came from a specific reel rather than the profile.
CURATED_POST_URLS = {
    "lukesoutpost": "https://www.instagram.com/reels/Da0pTNZumCD/",
}


def build_curated_sentiments(creators: list[dict]) -> list[dict]:
    """Hand-picked creator quotes from social posts. Labelled 'Creator' like the
    rest of the grid; the handle links out to the source post/profile."""
    by_handle = {c["handle"].lower(): c.get("name") for c in creators}
    out = []
    for handle, quote, category, emoji, tone in CURATED_CREATOR_QUOTES:
        real_name = by_handle.get(handle.lower())
        out.append({
            # fall back to the @handle when the creator isn't in the database
            "name": real_name or ("@" + handle),
            "role": ROLE_CREATOR,
            "quote": quote,
            "category": category,
            "emoji": emoji,
            "tone": tone,
            "source": "creator",
            "postUrl": CURATED_POST_URLS.get(handle.lower(), f"https://www.instagram.com/{handle}"),
        })
    return out


def source_from(cm: str, who: str, role: str) -> str:
    """Two buckets: 'client' (brand/event-team side) vs 'creator' (creators,
    members, attendees, panelists, pedestrians). Uses the 'Client or member'
    column, falling back to keywords."""
    blob = " ".join(x for x in (cm, who, role) if x).lower()
    if "client" in blob or "event team" in blob:
        return "client"
    return "creator"


def build_sentiments() -> list[dict]:
    wb = load_workbook(SENT_XLSX, data_only=True)
    ws = wb.active
    out = []
    for r in range(3, ws.max_row + 1):
        who = ws.cell(r, 1).value
        text = ws.cell(r, 2).value
        typ = ws.cell(r, 3).value
        clientmember = ws.cell(r, 4).value
        who = (who or "").strip()
        text = (text or "").strip()
        cm = (str(clientmember).strip() if clientmember else "")
        if not text:
            continue
        # Only surface rows that match one of the 12 curated moment categories —
        # skip generic thanks / metrics reports so the grid stays clean.
        cat = categorize_sentiment(text)
        if cat is None:
            continue
        label, emoji, tone = cat
        name, role = clean_attribution(who)
        # if the row's Type provides more useful context, mix it in
        if not role and typ:
            role = str(typ).strip().rstrip()
        src = source_from(cm, who, role)
        out.append({
            "name": name,
            "role": normalize_role(name, role, src),
            "quote": shorten_quote(text),
            "category": label,
            "emoji": emoji,
            "tone": tone,
            "source": src,
        })
    for extra in EXTRA_SENTIMENTS:
        e = dict(extra)
        e["role"] = normalize_role(e.get("name"), e.get("role"), e.get("source"))
        out.append(e)
    return out


# ---- sentiment pulled from creator content (post captions) ----

_QUOTE_STRIP = re.compile(r"https?://\S+")


def caption_quote(caption: str) -> str | None:
    """Pull a short, clean opening line from a creator's caption to use as a
    sentiment quote. Rejects lines that still contain @mentions, #hashtags or
    links so the card reads like a real quote."""
    if not caption:
        return None
    t = _QUOTE_STRIP.sub("", caption)
    for line in (l.strip() for l in t.split("\n")):
        if not line:
            continue
        m = re.match(r"(.+?[.!?…])(?:\s|$)", line)
        s = (m.group(1) if m else line).strip()
        if "@" in s or "#" in s or "http" in s.lower():
            return None  # first content line is tag-heavy — skip this creator
        s = re.sub(r"\s+", " ", s).strip(" .")
        if not (18 <= len(s) <= 140) or not s[0].isalpha():
            return None
        return s
    return None


# Curated allow-list for the "CREATOR" sentiment cards — kept to these four by
# request (everything else pulled from captions was cut from the grid).
CREATOR_SENTIMENT_HANDLES = {"sammyjreacts", "wafellow", "coyjandreau", "caitlinchristinee"}

# Manual quote overrides for curated CREATOR cards — longer excerpts than the
# single-sentence auto-extractor would pull, specified verbatim by request.
CREATOR_QUOTE_OVERRIDES = {
    "caitlinchristinee": "Countdown to Spider-Man begins now! 🕷️🕸️\n\nWords can't describe how excited I am to see my favorite superhero on screen again 😭",
}


def build_creator_sentiments(creators: list[dict], posts: list[dict], limit: int = 12) -> list[dict]:
    """One quotable line per creator, drawn from their highest-engagement post."""
    by_creator: dict[str, dict] = {}
    for p in posts:
        if p["platform"] not in ("tiktok", "instagram"):
            continue
        q = caption_quote(p.get("caption"))
        if not q:
            continue
        eng = sum(v for v in (p.get("metrics") or {}).values() if isinstance(v, (int, float)))
        key = p["creatorHandle"].lower()
        if key not in by_creator or eng > by_creator[key]["_eng"]:
            by_creator[key] = {"quote": q, "handle": p["creatorHandle"], "platform": p["platform"],
                               "postUrl": p["postUrl"], "_eng": eng}
    cmap = {c["handle"].lower(): c for c in creators}
    by_creator = {k: v for k, v in by_creator.items() if k in CREATOR_SENTIMENT_HANDLES}
    items = sorted(by_creator.values(), key=lambda x: x["_eng"], reverse=True)[:limit]
    out = []
    for it in items:
        c = cmap.get(it["handle"].lower(), {})
        quote = CREATOR_QUOTE_OVERRIDES.get(it["handle"].lower(), it["quote"])
        out.append({
            "name": c.get("name") or it["handle"],
            "role": ("TikTok" if it["platform"] == "tiktok" else "Instagram") + " · from their post",
            "quote": quote,
            "category": "CREATOR",
            "emoji": "🕷️",
            "tone": "teal",
            "source": "creator",
            "handle": it["handle"],
            "postUrl": it["postUrl"],
        })
    return out


# Additional real feedback from second.xlsx that the keyword pass skips because
# it doesn't map to one of the punchy moment categories. Curated here (verbatim
# quotes, cleaned only for length) so "What People Said" isn't so sparse.
EXTRA_SENTIMENTS = [
    {"name": "Maria", "role": "Client",
     "quote": "Thank you so much Lauren — such a pleasure working with you and the whole Lighthouse team!",
     "category": "GRATEFUL", "emoji": "🙏", "tone": "pink", "source": "client"},
    {"name": "Jenna", "role": "Event Team",
     "quote": "We truly cannot thank you enough for everything. It was such a genuine pleasure to work at your space and with each of you. What a special event!",
     "category": "GRATEFUL", "emoji": "🙏", "tone": "pink", "source": "client"},
    {"name": "Geo", "role": "Client",
     "quote": "General Event Photos are all GREAT!",
     "category": "GLOWING", "emoji": "✨", "tone": "teal", "source": "client"},
    {"name": "Geo", "role": "Client",
     "quote": "The RSVPs are looking great — thanks to everyone who contributed to the amazing attendance!",
     "category": "BUZZING", "emoji": "📈", "tone": "teal", "source": "client"},
]


# ---- Owned & collaborative posts (The Lighthouse's own + official Spider-Man collab) ----
# Metadata is curated here; thumbnails + account avatars come from owned_images.json
# (fetched from the public post pages). Metrics are the public like/comment counts.
OWNED_POSTS = [
    {"shortcode": "DbbmXcBSMN5", "url": "https://www.instagram.com/p/DbbmXcBSMN5/",
     "account": "spidermanmovie", "accountName": "Spider-Man", "official": True, "event": EVENT_VENICE,
     "likes": "13K", "comments": "173",
     "caption": "“This is our time to connect.” — @destindaniel. A recap of the @tiktokcreators + "
                "@thelighthousecampus event with the Spider-Man: Brand New Day filmmakers.",
     "collab": ["thelighthousecampus", "tiktokcreators"]},
    {"shortcode": "DbmCrdgPVPj", "url": "https://www.instagram.com/p/DbmCrdgPVPj/",
     "account": "thelighthousecampus", "accountName": "The Lighthouse", "official": False, "event": EVENT_VENICE,
     "likes": "14K", "comments": "41",
     "caption": "Views from The Lighthouse Venice — the Spider-Man: Brand New Day Creator × Filmmaker "
                "Experience. 200+ creators, filmmakers, and entertainment voices on our Venice campus.",
     "collab": ["sonypictures", "tiktokcreators"]},
    {"shortcode": "DbYnq_VKlgG", "url": "https://www.instagram.com/p/DbYnq_VKlgG/",
     "account": "thelighthousecampus", "accountName": "The Lighthouse", "official": False, "event": EVENT_VENICE,
     "likes": None, "comments": None,
     "caption": "It’s a Brand New Day at The Lighthouse. We joined Lighthouse Creator @lonnieiiv for the "
                "@sonypictures Spider-Man: Brand New Day Creator Filmmaker Experience in Venice.",
     "collab": ["lonnieiiv", "sonypictures"]},
    {"shortcode": "DbZv-9vt7v3", "url": "https://www.instagram.com/p/DbZv-9vt7v3/",
     "account": "thelighthousecampus", "accountName": "The Lighthouse", "official": False, "event": EVENT_BROOKLYN,
     "likes": "1,229", "comments": "36",
     "caption": "A Brand New Day landed at The Lighthouse Brooklyn. @sonypictures and The Lighthouse came "
                "together for a special early screening — 100+ creators in our theater.",
     "collab": ["sonypictures"]},
]


def load_owned() -> list[dict]:
    imgs = {"thumbs": {}, "accounts": {}}
    f = HERE / "owned_images.json"
    if f.exists():
        imgs = json.loads(f.read_text())
    out = []
    for p in OWNED_POSTS:
        q = dict(p)
        q["thumb"] = imgs.get("thumbs", {}).get(p["shortcode"])
        q["accountAvatar"] = imgs.get("accounts", {}).get(p["account"])
        out.append(q)
    return out


def main():
    creators, posts, stats = build_posts()
    # Sentiment cards are drawn from the Venice event only.
    sentiments = build_sentiments()
    sentiments += build_creator_sentiments(creators, posts)
    sentiments += build_curated_sentiments(creators)

    bk_creators, bk_posts, bk_stats = build_brooklyn()
    if bk_posts:
        print(f"brooklyn: {len(bk_creators)} creators, {len(bk_posts)} instagram posts")
        creators = creators + bk_creators
        posts = posts + bk_posts

    # Quotes to drop from the sentiment grid (matched by a lowercase substring).
    SENTIMENT_EXCLUDE = (
        "wait is that real",           # Vidisha Jain — removed by request
        "nothing ever happens out here",  # The Lighthouse Member — removed by request
    )

    # De-duplicate sentiments by quote text (some feedback repeats) and drop excludes
    seen_q = set()
    uniq_sentiments = []
    for s in sentiments:
        key = s["quote"].lower()
        if key in seen_q or any(x in key for x in SENTIMENT_EXCLUDE):
            continue
        seen_q.add(key)
        uniq_sentiments.append(s)

    # Combined (both events) totals — the engagement panel recomputes per event
    # at runtime, but these keep the build output and any static fallbacks honest.
    combined_eng = {k: stats["engagement"][k] + bk_stats.get("engagement", {}).get(k, 0)
                    for k in ("likes", "comments", "savesReposts", "shares")}
    all_stats = {
        "creators": len(creators),
        "combinedFollowing": stats["combinedFollowing"] + bk_stats.get("combinedFollowing", 0),
        "millionPlus": stats["millionPlus"] + bk_stats.get("millionPlus", 0),
        "postsLogged": stats["postsLogged"] + bk_stats.get("postsLogged", 0),
        "tiktokPosts": stats["tiktokPosts"],
        "instagramPosts": stats["instagramPosts"] + bk_stats.get("instagramPosts", 0),
        "engagement": combined_eng,
        "totalEngagement": sum(combined_eng.values()),
    }

    payload = {
        "creators": creators,
        "posts": posts,
        "owned": load_owned(),
        "sentiments": uniq_sentiments,
        "stats": all_stats,
        "events": {
            EVENT_VENICE: {"label": "Venice Event",
                           "name": "Sony Pictures × TikTok — Venice",
                           "stats": stats},
            EVENT_BROOKLYN: {"label": "Brooklyn Event",
                             "name": "Brooklyn Advance Creator Screening",
                             "stats": bk_stats},
        },
        "counts": {
            "posts": all_stats["postsLogged"],
            "creators": len(creators),
            "sentiments": len(uniq_sentiments),
            "tiktok": all_stats["tiktokPosts"],
            "instagram": all_stats["instagramPosts"],
            "stories": sum(1 for p in posts if p["platform"] == "story"),
        },
    }

    blob = "/*__DATA_START__*/\nwindow.__RECAP_DATA__ = " + json.dumps(payload, ensure_ascii=False) + ";\n/*__DATA_END__*/"

    if not INDEX_HTML.exists():
        print(f"NOTE: {INDEX_HTML} not found yet — writing data blob to data.js instead")
        (HERE / "data.js").write_text(blob, encoding="utf-8")
    else:
        html = INDEX_HTML.read_text(encoding="utf-8")
        pattern = re.compile(r"/\*__DATA_START__\*/[\s\S]*?/\*__DATA_END__\*/")
        if not pattern.search(html):
            raise SystemExit("index.html is missing the /*__DATA_START__*/.../*__DATA_END__*/ sentinels")
        new_html = pattern.sub(lambda _m: blob, html)

        # Inline baked-in avatars (from fetch_avatars.py) into the __AVATARS block.
        avatars_file = HERE / "avatars.json"
        if avatars_file.exists():
            avatars = json.loads(avatars_file.read_text())
            av_blob = ("/*__AVATARS_START__*/\nwindow.__AVATARS = "
                       + json.dumps(avatars, ensure_ascii=False) + ";\n/*__AVATARS_END__*/")
            av_pat = re.compile(r"/\*__AVATARS_START__\*/[\s\S]*?/\*__AVATARS_END__\*/")
            if av_pat.search(new_html):
                new_html = av_pat.sub(lambda _m: av_blob, new_html)
                print(f"avatars: inlined {len(avatars)} baked-in profile images")

        INDEX_HTML.write_text(new_html, encoding="utf-8")

    print(f"creators: {len(creators)}")
    print(f"posts:    {payload['counts']['posts']}  (tt={payload['counts']['tiktok']}, ig={payload['counts']['instagram']}, stories={payload['counts']['stories']})")
    print(f"sentiments: {len(uniq_sentiments)}")
    print(f"stats: checkedIn={stats['checkedIn']} following={stats['combinedFollowing']:,} 1M+={stats['millionPlus']} totalEng={stats['totalEngagement']:,}")

    # Top mentions/hashtags for sanity
    from collections import Counter
    mc, hc = Counter(), Counter()
    for p in posts:
        for m in p["mentions"]:
            mc[m["norm"]] += 1
        for h in p["hashtags"]:
            hc[h.lower()] += 1
    print(f"top mentions: {mc.most_common(10)}")
    print(f"top hashtags: {hc.most_common(10)}")


if __name__ == "__main__":
    main()
